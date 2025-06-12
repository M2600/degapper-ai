# 渡されたデータをexcel形式で保存する
import argparse
import json

from collections import defaultdict
from pprint import pprint

import openpyxl as pxl

from tools.utils import Input


def getLabels(ipt):
    return ipt[0].split(",")


def writeLabels(sheet, labels):
    for i, label in enumerate(labels, start=1):
        sheet.cell(row=1, column=i).value = label


def writeRow(sheet, data, start):
    for i, d in enumerate(data, start=1):
        try:
            d = float(d)
        except:
            pass
        sheet.cell(row=start, column=i).value = d


def writeWB(times, labels):
    # フォルダごとのシートを作って書き込み
    wb = pxl.Workbook()
    wb.remove(wb.worksheets[0])
    for time, data in times.items():
        sheet = wb.create_sheet(time)
        writeLabels(sheet, labels)
        for i, d in enumerate(data, start=2):
            writeRow(sheet, d, i)
    return wb


def saveParseTree(ipt):
    labels = getLabels(ipt)

    # フォルダごとに分ける
    times = defaultdict(list)  # {"1st": [file, file]}
    for row in ipt[1:]:
        name, route = row.split(",")
        time, id = name.split("\\")[-2:]
        times[time].append([id, route])

    # フォルダごとのシートを作って書き込み
    wb = writeWB(times, labels)

    folder = ipt[1].split(",")[0].split("\\")[0]
    wb.save(f"{folder}/parse_tree.xlsx")


def saveDegapper(ipt):
    labels = getLabels(ipt)

    # フォルダごとに分ける
    times = defaultdict(list)  # {"time": [[name, gap, m/m]]}
    for row in ipt[1:]:
        name, gap, mm = row.split(",")
        time, id = name.split("\\")[-2:]
        times[time].append([id, gap, mm])

    # フォルダごとのシートを作って書き込み
    wb = writeWB(times, labels)

    folder = ipt[1].split(",")[0].split("\\")[0]
    wb.save(f"{folder}/degapper.xlsx")


def saveCountMajorMinor(ipt):
    labels = getLabels(ipt)

    # フォルダごとに分ける
    times = defaultdict(list)  # {"time": [[name, gap, m/m]]}
    for row in ipt[1:]:
        name, gap, mm = row.split(",")
        time, id = name.split("\\")[-2:]
        times[time].append([id, gap, mm])

    # フォルダごとのシートを作って書き込み
    wb = writeWB(times, labels)

    folder = ipt[1].split(",")[0].split("\\")[0]
    wb.save(f"{folder}/count_major_minor.xlsx")


def saveCalcMetrics(ipt):
    labels = getLabels(ipt)

    # file, loc, cc, hv, mi

    # フォルダごとに分ける
    times = defaultdict(list)  # {"time": [[name, loc, cc, hv, mi]]}
    for row in ipt[1:]:
        name, *metrics = row.split(",")
        time, id = name.split("\\")[-2:]
        times[time].append([id, *metrics])

    # フォルダごとのシートを作って書き込み
    wb = writeWB(times, labels)

    folder = ipt[1].split(",")[0].split("\\")[0]
    wb.save(f"{folder}/code_metrics.xlsx")


def saveComponents(ipt):
    labels = getLabels(ipt)

    # フォルダごとに分ける
    times = defaultdict(list)  # {time: [[name, components]]}
    for row in ipt[1:]:
        name, *components = row.split(",")
        time, fileName = name.split("/")[-2:]
        times[time].append([fileName, *components])

    # フォルダごとのシートを作って書き込み
    wb = writeWB(times, labels)

    res = ipt[1].split(",")[0].split("/")
    folder = "/".join(res[:-2])
    wb.save(f"{folder}/components.xlsx")


def saveNestedStructure(ipt):
    # 出力形式が他と違うので、個別に実装
    nested = json.loads("\n".join(ipt))  # {folder: file: [max_nest, structures]}

    wb = pxl.Workbook()
    wb.remove(wb.worksheets[0])
    for folder, data in nested.items():
        folder = folder.split("/")
        folder = folder[-1] if folder[-1] != "" else folder[-2]

        sheet = wb.create_sheet(folder)
        sheet.cell(row=1, column=1).value = "file"
        sheet.cell(row=1, column=2).value = "max_nest_size"
        sheet.cell(row=1, column=3).value = "nest_structure"
        for row, (file, (maxNest, structures)) in enumerate(data.items(), start=2):
            sheet.cell(row=row, column=1).value = file
            sheet.cell(row=row, column=2).value = int(maxNest)

            # ネスト構造が無ければ「無し」
            if len(structures) == 0:
                sheet.cell(row=row, column=3).value = "無し"
                continue

            # 検出したネスト構造を列挙する
            for col, elem in enumerate(structures, start=3):
                sheet.cell(row=row, column=col).value = elem
    saveFolder = "/".join(list(nested.keys())[0].split("/")[:-2])
    wb.save(f"{saveFolder}/nest_structure.xlsx")


ipt = Input()

wb = pxl.Workbook()
sheet = wb.worksheets[0]

try:
    if ipt[0] == "name,component":
        saveParseTree(ipt)
    elif ipt[0] == "name,gap,major/minor":
        saveDegapper(ipt)
    elif ipt[0] == "file,major,minor":
        saveCountMajorMinor(ipt)
    elif ipt[0] == "file,loc,cc,hv,mi":
        saveCalcMetrics(ipt)
    elif ipt[0] == "file,output,variable,input,if/else,for,while,array,nest,function,and/or,elseif,loop,dict":
        saveComponents(ipt)
    elif ipt[0] == "{":  # json形式
        saveNestedStructure(ipt)
    else:
        print("この出力はexcel出力に対応していません")
except PermissionError:
    print("エクセルファイルを閉じてください")
